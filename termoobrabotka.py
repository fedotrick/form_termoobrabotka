import sys
import os
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLineEdit,
    QPushButton, QMessageBox, QLabel, QComboBox, QDateEdit, QHBoxLayout
)
from PySide6.QtCore import Qt, QDate, QTimer
from openpyxl import Workbook, load_workbook
from PySide6.QtGui import QPalette, QColor, QFont

# Функция для сохранения данных в Excel
def save_to_excel(номер_плавки, термообработка_номер_печи, термообработка_дата,
                 термообработка_начало_первого_цикла, термообработка_конец_первого_цикла,
                 термообработка_начало_второго_цикла="", термообработка_конец_второго_цикла=""):
    wb = None
    try:
        headers = ['Номер плавки', 'Номер печи', 'Дата', 
                  'Начало первого цикла', 'Конец первого цикла',
                  'Начало второго цикла', 'Конец второго цикла']

        # Если файл не существует, создаем его с заголовками
        if not os.path.exists('termoobrabotka.xlsx'):
            wb = Workbook()
            ws = wb.active
            ws.title = "Records"
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)
            row_to_write = 2  # Первая строка для данных
        else:
            wb = load_workbook('termoobrabotka.xlsx')
            if "Records" not in wb.sheetnames:
                ws = wb.create_sheet("Records")
                # Добавляем заголовки для нового листа
                for col, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col, value=header)
                row_to_write = 2
            else:
                ws = wb["Records"]
                # Проверяем, существует ли уже запись с таким номером плавки
                existing_row = None
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value == номер_плавки:
                        existing_row = row
                        break
                
                if existing_row:
                    row_to_write = existing_row
                else:
                    row_to_write = ws.max_row + 1

        # Записываем значения
        values = [номер_плавки, термообработка_номер_печи, термообработка_дата,
                 термообработка_начало_первого_цикла, термообработка_конец_первого_цикла,
                 термообработка_начало_второго_цикла, термообработка_конец_второго_цикла]
        
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_to_write, column=col)
            cell.value = value
            
            if col == 3:  # Колонка C (дата)
                cell.number_format = 'DD.MM.YYYY'
            elif col in [4, 5, 6, 7]:  # Колонки D, E, F, G (время)
                cell.number_format = 'HH:MM'
        
        wb.save('termoobrabotka.xlsx')
    except Exception as e:
        raise Exception(f"Ошибка при сохранении в Excel: {str(e)}")
    finally:
        if wb:
            wb.close()

def get_existing_plavki():
    file_name = 'plavka.xlsx'
    if not os.path.exists(file_name):
        return []

    wb = None
    try:
        wb = load_workbook(file_name)
        sheet = wb.active
        номера_плавок = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            номер_плавки = row[1]  
            if номер_плавки is not None:
                номера_плавок.append(str(номер_плавки))

        return номера_плавок
    finally:
        if wb:
            wb.close()

def get_available_plavki():
    # Получаем все номера плавок из plavka.xlsx
    все_плавки = get_existing_plavki()
    
    # Получаем номера плавок, которые уже существуют в termoobrabotka.xlsx
    существующие_плавки = set()
    file_name = 'termoobrabotka.xlsx'
    wb = None
    try:
        if os.path.exists(file_name):
            wb = load_workbook(file_name)
            if "Records" in wb.sheetnames:
                sheet = wb["Records"]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    существующие_плавки.add(row[0])  # Первый столбец содержит номер плавки

        # Фильтруем номера плавок
        доступные_плавки = [
            плавка for плавка in все_плавки 
            if плавка not in существующие_плавки and '/25' in плавка
        ]

        # Сортируем по убыванию
        доступные_плавки.sort(reverse=True)

        return доступные_плавки
    finally:
        if wb:
            wb.close()


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.plавка_fields = []
        
        self.setWindowTitle("Электронный журнал термообработки")
        self.setMinimumWidth(600)
        self.setMinimumHeight(400)
        
        # Обновляем стили для лучшей читаемости
        self.setStyleSheet("""
            QWidget {
                background-color: #2B2B2B;
                color: #E6E6E6;
                font-family: 'Segoe UI', 'Arial';
                font-size: 12px;
            }
            QLabel {
                color: #FF6B35;
                font-size: 13px;
                font-weight: bold;
                padding: 8px;
                border: 2px solid #4A4A4A;
                border-radius: 5px;
                background-color: #1A1A1A;
                margin: 2px;
                text-transform: uppercase;
                letter-spacing: 1px;
            }
            QComboBox {
                background-color: #333333;
                border: 2px solid #4A4A4A;
                border-radius: 5px;
                padding: 5px;
                color: #FFFFFF;
                min-height: 25px;
                margin: 2px;
                font-size: 12px;
                font-weight: bold;
            }
            QComboBox::drop-down {
                border: none;
                width: 25px;
                background-color: #4A4A4A;
            }
            QComboBox:hover {
                background-color: #3D3D3D;
                border-color: #FF6B35;
            }
            QLineEdit {
                background-color: #333333;
                border: 2px solid #4A4A4A;
                border-radius: 5px;
                padding: 5px;
                color: #FFFFFF;
                min-height: 25px;
                margin: 2px;
                font-size: 12px;
                font-weight: bold;
            }
            QLineEdit:focus {
                border-color: #FF6B35;
            }
            QDateEdit {
                background-color: #333333;
                border: 2px solid #4A4A4A;
                border-radius: 5px;
                padding: 5px;
                color: #FFFFFF;
                min-height: 25px;
                margin: 2px;
                font-size: 12px;
                font-weight: bold;
            }
            QDateEdit::drop-down {
                border: none;
                width: 25px;
                background-color: #4A4A4A;
            }
            QPushButton {
                background-color: #FF6B35;
                color: #FFFFFF;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font-size: 14px;
                font-weight: bold;
                min-height: 30px;
                margin: 5px;
                text-transform: uppercase;
                letter-spacing: 2px;
            }
            QPushButton:hover {
                background-color: #FF8C61;
            }
            QPushButton:pressed {
                background-color: #E65A2C;
                padding: 9px 14px 7px 16px;
            }
            
            /* Стиль для заголовка */
            QLabel#title {
                font-size: 22px;
                font-weight: bold;
                color: #FFFFFF;
                border: 3px solid #FF6B35;
                padding: 10px;
                background-color: #1A1A1A;
                margin: 5px;
                letter-spacing: 3px;
            }
        """)

        layout = QVBoxLayout()
        layout.setSpacing(2)  # Уменьшаем расстояние между элементами
        layout.setContentsMargins(5, 5, 5, 5)  # Уменьшаем отступы от краёв

        # Обновляем стиль заголовка
        title = QLabel("ЭЛЕКТРОННЫЙ ЖУРНАЛ ТЕРМООБРАБОТКИ")
        title.setObjectName("title")  # Добавляем id для стилизации
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Создаем горизонтальные контейнеры для группировки элементов
        top_container = QWidget()
        top_layout = QHBoxLayout(top_container)
        top_layout.setSpacing(2)
        top_layout.setContentsMargins(0, 0, 0, 0)

        # Левая часть с печью и плавками
        left_container = QWidget()
        left_layout = QVBoxLayout(left_container)
        left_layout.setSpacing(2)
        left_layout.setContentsMargins(0, 0, 0, 0)

        печь_label = QLabel("НОМЕР ПЕЧИ")
        печь_label.setAlignment(Qt.AlignCenter)
        left_layout.addWidget(печь_label)

        self.термообработка_номер_печи = QComboBox()
        self.термообработка_номер_печи.addItems(['1', '2'])
        self.термообработка_номер_печи.currentTextChanged.connect(self.update_plavka_fields)
        left_layout.addWidget(self.термообработка_номер_печи)

        # Контейнер для плавок
        self.plавка_container = QWidget()
        self.plавка_layout = QVBoxLayout(self.plавка_container)
        self.plавка_layout.setSpacing(2)
        self.plавка_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.addWidget(self.plавка_container)

        top_layout.addWidget(left_container)

        # Правая часть с датой и циклами
        right_container = QWidget()
        right_layout = QVBoxLayout(right_container)
        right_layout.setSpacing(2)
        right_layout.setContentsMargins(0, 0, 0, 0)

        дата_label = QLabel("ТЕРМООБРАБОТКА")
        дата_label.setAlignment(Qt.AlignCenter)
        дата_label.setObjectName("termo_label")  # Добавляем id для стилизации
        дата_label.setStyleSheet("""
            QLabel#termo_label {
                color: #FFFFFF;
                font-size: 32px;
                font-weight: bold;
                padding: 5px;
                border: 1px solid #00ffff;
                border-radius: 3px;
                background-color: #2a2a2a;
                margin: 1px;
                background-image: url(termo.png);
                background-position: center;
                background-repeat: no-repeat;
                background-origin: content;
            }
        """)
        right_layout.addWidget(дата_label)

        # Сохраняем ссылку на метку как атрибут класса
        self.termo_label = дата_label

        # Создаем таймер для смены цветов
        self.color_timer = QTimer(self)
        self.color_timer.timeout.connect(self.update_label_color)
        self.color_timer.start(1000)  # Интервал в миллисекундах (1000 = 1 секунда)

        # Список цветов для анимации
        self.colors = [
            "#FF0000",  # Красный
            "#FF7F00",  # Оранжевый
            "#FFFF00",  # Желтый
            "#00FF00",  # Зеленый
            "#0000FF",  # Синий
            "#4B0082",  # Индиго
            "#9400D3"   # Фиолетовый
        ]
        self.current_color_index = 0

        self.термообработка_дата = QDateEdit()
        self.термообработка_дата.setDisplayFormat("dd.MM.yyyy")
        self.термообработка_дата.setCalendarPopup(True)
        self.термообработка_дата.setDate(QDate.currentDate())
        right_layout.addWidget(self.термообработка_дата)

        # Контейнер для циклов
        cycles_container = QWidget()
        cycles_layout = QHBoxLayout(cycles_container)
        cycles_layout.setSpacing(2)
        cycles_layout.setContentsMargins(0, 0, 0, 0)

        # Первый цикл
        cycle1_container = QWidget()
        cycle1_layout = QVBoxLayout(cycle1_container)
        cycle1_layout.setSpacing(2)
        cycle1_layout.setContentsMargins(0, 0, 0, 0)

        цикл1_label = QLabel("ПЕРВЫЙ ЦИКЛ")
        цикл1_label.setAlignment(Qt.AlignCenter)
        cycle1_layout.addWidget(цикл1_label)

        self.термообработка_начало_первого_цикла = QLineEdit()
        self.термообработка_начало_первого_цикла.setPlaceholderText("Начало (ЧЧ:ММ)")
        self.термообработка_начало_первого_цикла.setMaxLength(5)
        self.термообработка_начало_первого_цикла.textChanged.connect(self.format_time_input)
        cycle1_layout.addWidget(self.термообработка_начало_первого_цикла)

        self.термообработка_конец_первого_цикла = QLineEdit()
        self.термообработка_конец_первого_цикла.setPlaceholderText("Конец (ЧЧ:ММ)")
        self.термообработка_конец_первого_цикла.setMaxLength(5)
        self.термообработка_конец_первого_цикла.textChanged.connect(self.format_time_input)
        cycle1_layout.addWidget(self.термообработка_конец_первого_цикла)

        cycles_layout.addWidget(cycle1_container)

        # Второй цикл
        cycle2_container = QWidget()
        cycle2_layout = QVBoxLayout(cycle2_container)
        cycle2_layout.setSpacing(2)
        cycle2_layout.setContentsMargins(0, 0, 0, 0)

        цикл2_label = QLabel("ВТОРОЙ ЦИКЛ")
        цикл2_label.setAlignment(Qt.AlignCenter)
        cycle2_layout.addWidget(цикл2_label)

        self.термообработка_начало_второго_цикла = QLineEdit()
        self.термообработка_начало_второго_цикла.setPlaceholderText("Начало (ЧЧ:ММ)")
        self.термообработка_начало_второго_цикла.setMaxLength(5)
        self.термообработка_начало_второго_цикла.textChanged.connect(self.format_time_input)
        cycle2_layout.addWidget(self.термообработка_начало_второго_цикла)

        self.термообработка_конец_второго_цикла = QLineEdit()
        self.термообработка_конец_второго_цикла.setPlaceholderText("Конец (ЧЧ:ММ)")
        self.термообработка_конец_второго_цикла.setMaxLength(5)
        self.термообработка_конец_второго_цикла.textChanged.connect(self.format_time_input)
        cycle2_layout.addWidget(self.термообработка_конец_второго_цикла)

        cycles_layout.addWidget(cycle2_container)
        right_layout.addWidget(cycles_container)

        top_layout.addWidget(right_container)
        layout.addWidget(top_container)

        # Кнопка сохранения
        self.save_button = QPushButton("СОХРАНИТЬ")
        self.save_button.clicked.connect(self.save_data)
        layout.addWidget(self.save_button)

        self.setLayout(layout)
        
        # Инициализация полей плавок
        self.selected_plavki = set()
        self.update_plavka_fields('1')

    def update_plavka_fields(self, печь_номер):
        # Сбрасываем выбранные плавки при обновлении полей
        self.selected_plavki.clear()
        
        # Очищаем существующие поля
        for field in self.plавка_fields:
            self.plавка_layout.removeWidget(field)
            field.deleteLater()
        self.plавка_fields.clear()

        # Определяем количество полей в зависимости от номера печи
        количество_полей = 10 if печь_номер == '1' else 9
        
        # Получаем доступные плавки
        доступные_плавки = get_available_plavki()
        
        # Создаем новые поля
        for i in range(количество_полей):
            combo = QComboBox()
            combo.setObjectName(f"plавка_{i}")  # Добавляем уникальное имя
            combo.addItem(f"ПЛАВКА {i+1}")
            combo.addItems(доступные_плавки)
            combo.currentTextChanged.connect(self.on_plavka_selected)
            self.plавка_fields.append(combo)
            self.plавка_layout.addWidget(combo)

    def on_plavka_selected(self, selected_value):
        sender = self.sender()
        previous_value = sender.property("previous_value")
        
        # Проверяем, действительно ли значение изменилось
        if selected_value == previous_value:
            return
        
        # Удаляем предыдущее значение из selected_plavki, если оно было
        if previous_value and not previous_value.startswith("ПЛАВКА"):
            self.selected_plavki.discard(previous_value)
            
        # Добавляем новое значение, если это не заголовок
        if not selected_value.startswith("ПЛАВКА"):
            self.selected_plavki.add(selected_value)
            
        # Сохраняем текущее значение как предыдущее
        sender.setProperty("previous_value", selected_value)
        
        # Обновляем доступные плавки во всех комбобоксах
        доступные_плавки = get_available_plavki()
        
        # Обновляем каждый комбобокс
        for combo in self.plавка_fields:
            if combo != sender:
                current_value = combo.currentText()
                combo.blockSignals(True)  # Блокируем сигналы во время обновления
                combo.clear()
                
                # Добавляем заголовок
                combo.addItem(f"ПЛАВКА {self.plавка_fields.index(combo)+1}")
                
                # Добавляем доступные плавки
                for плавка in доступные_плавки:
                    if плавка not in self.selected_plavki or плавка == current_value:
                        combo.addItem(плавка)
                
                # Восстанавливаем текущее значение если оно еще доступно
                index = combo.findText(current_value)
                if index >= 0:
                    combo.setCurrentIndex(index)
                else:
                    combo.setCurrentIndex(0)  # Устанавливаем на заголовок если значение недоступно
                    
                combo.blockSignals(False)  # Разблокируем сигналы

    def format_time_input(self, text):
        """Автоматически добавляет двоеточие после двух цифр"""
        if len(text) == 2 and text.isdigit():
            # Проверяем, что часы в допустимом диапазоне
            if 0 <= int(text) < 24:
                self.sender().setText(text + ":")
                self.sender().setCursorPosition(3)
            else:
                self.sender().setText("23:")
                self.sender().setCursorPosition(3)

    def validate_time(self, time_str):
        """Проверка корректности ввода времени в формате ЧЧ:ММ"""
        try:
            hours, minutes = map(int, time_str.split(':'))
            if 0 <= hours < 24 and 0 <= minutes < 60:
                return True
        except ValueError:
            return False
        return False

    def save_data(self):
        # Проверка на дублирование плавок
        выбранные_плавки = [
            combo.currentText() for combo in self.plавка_fields 
            if not combo.currentText().startswith("ПЛАВКА")
        ]
        
        if not выбранные_плавки:
            QMessageBox.warning(self, "Ошибка", "Выберите хотя бы одну плавку.")
            return
        
        if len(set(выбранные_плавки)) != len(выбранные_плавки):
            QMessageBox.warning(self, "Ошибка", "Обнаружено дублирование плавок.")
            return
        
        номер_печи = self.термообработка_номер_печи.currentText()
        термообработка_дата = self.термообработка_дата.date().toString("dd.MM.yyyy")
        начало_первого_цикла = self.термообработка_начало_первого_цикла.text().strip()
        конец_первого_цикла = self.термообработка_конец_первого_цикла.text().strip()
        начало_второго_цикла = self.термообработка_начало_второго_цикла.text().strip()
        конец_второго_цикла = self.термообработка_конец_второго_цикла.text().strip()

        # Проверка времени для первого цикла
        if not self.validate_time(начало_первого_цикла) or not self.validate_time(конец_первого_цикла):
            QMessageBox.warning(self, "Ошибка", "Некорректный ввод времени первого цикла. Используйте формат ЧЧ:ММ.")
            return

        # Проверка времени для второго цикла
        if (начало_второго_цикла and not конец_второго_цикла) or (not начало_второго_цикла and конец_второго_цикла):
            QMessageBox.warning(self, "Ошибка", "Для второго цикла должны быть заполнены оба времени либо оба пустые.")
            return

        if начало_второго_цикла and конец_второго_цикла:
            if not self.validate_time(начало_второго_цикла) or not self.validate_time(конец_второго_цикла):
                QMessageBox.warning(self, "Ошибка", "Некорректный ввод времени второго цикла. Используйте формат ЧЧ:ММ.")
                return

        try:
            # Сохраняем данные только для реально выбранных плавок
            for плавка in выбранные_плавки:
                save_to_excel(
                    плавка,
                    номер_печи,
                    термообработка_дата,
                    начало_первого_цикла,
                    конец_первого_цикла,
                    начало_второго_цикла,
                    конец_второго_цикла
                )

            QMessageBox.information(self, "Успех", "Данные сохранены в Excel!")
            self.clear_fields()
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при сохранении данных: {str(e)}")

    def clear_fields(self):
        self.selected_plavki.clear()
        self.термообработка_дата.setDate(QDate.currentDate())
        self.термообработка_начало_первого_цикла.clear()
        self.термообработка_конец_первого_цикла.clear()
        self.термообработка_начало_второго_цикла.clear()
        self.термообработка_конец_второго_цикла.clear()
        
        # Обновляем список доступных плавок после очистки
        self.update_plavka_fields(self.термообработка_номер_печи.currentText())

    def is_plavka_available(self, плавка, current_combo=None):
        """
        Проверяет, доступна ли плавка для выбора
        Args:
            плавка: номер плавки для проверки
            current_combo: текущий комбобокс (чтобы не исключать его текущее значение)
        """
        if плавка.startswith("ПЛАВКА"):
            return True
        if current_combo and current_combo.currentText() == плавка:
            return True
        return плавка not in self.selected_plavki

    def update_label_color(self):
        """Обновляет цвет метки термообработки"""
        color = self.colors[self.current_color_index]
        self.termo_label.setStyleSheet(f"""
            QLabel#termo_label {{
                color: {color};
                font-size: 32px;
                font-weight: bold;
                padding: 5px;
                border: 1px solid {color};
                border-radius: 3px;
                background-color: #2a2a2a;
                margin: 1px;
                background-image: url(termo.png);
                background-position: center;
                background-repeat: no-repeat;
                background-origin: content;
            }}
        """)
        
        # Переходим к следующему цвету
        self.current_color_index = (self.current_color_index + 1) % len(self.colors)

    def closeEvent(self, event):
        """Останавливаем таймер при закрытии окна"""
        self.color_timer.stop()
        super().closeEvent(event)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
